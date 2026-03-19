"""
报告生成器模块
功能: 生成 Excel 报告
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, Union

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


logger = logging.getLogger("paperinsight.reporter")


class ReportGenerator:
    """报告生成器 v3.0
    
    按照 PRD 3.0 规范生成 Excel 报告：
    - 主信息列（标题、作者、IF）独占单元格
    - 多器件数据列（结构、EQE、CIE、寿命）使用 \n 换行拼接
    """

    REPORT_COLUMNS = [
        ("文件名 File", "file"),
        ("文件地址 URL", "url"),
        ("期刊名称 Journal", "journal"),
        ("影响因子 Impact Factor", "impact_factor"),
        ("作者 Authors", "authors"),
        ("处理结果/简述 Processing Status", "processing_status"),
        ("论文标题 Title", "title"),
        ("器件结构 Device Structure", "structure"),
        ("EQE(外量子效率) EQE", "eqe"),
        ("色度坐标 CIE", "cie"),
        ("寿命 Lifetime", "lifetime"),
        ("最高EQE Best EQE", "best_eqe"),
        ("优化层级 Optimization Level", "optimization_level"),
        ("优化策略 Strategy Summary", "optimization_strategy"),
        ("优化详情 Optimization Details", "optimization_details"),
        ("关键发现 Key Findings", "key_findings"),
        ("EQE原文 EQE Source", "eqe_source"),
        ("CIE原文 CIE Source", "cie_source"),
        ("寿命原文 Lifetime Source", "lifetime_source"),
        ("结构原文 Structure Source", "structure_source"),
    ]

    FIELD_MAPPING = {
        "file": ("File", "file", "文件", "文件名"),
        "url": ("URL", "url", "文件地址"),
        "journal": ("journal_name", "期刊", "期刊名称"),
        "raw_journal_title": ("raw_journal_title", "原始期刊标题"),
        "raw_issn": ("raw_issn", "原始ISSN"),
        "raw_eissn": ("raw_eissn", "原始eISSN"),
        "matched_journal_title": ("matched_journal_title", "匹配期刊"),
        "matched_issn": ("matched_issn", "匹配ISSN"),
        "match_method": ("match_method", "匹配方式"),
        "journal_profile_url": ("journal_profile_url", "期刊主页"),
        "impact_factor": ("影响因子", "impact_factor"),
        "impact_factor_year": ("impact_factor_year", "影响因子年份"),
        "impact_factor_source": ("impact_factor_source", "影响因子来源"),
        "impact_factor_status": ("impact_factor_status", "影响因子状态"),
        "authors": ("authors", "作者"),
        "processing_status": ("processing_status", "处理结果/简述", "处理结果", "简述"),
        "title": ("title", "标题", "论文标题"),
        "structure": ("器件结构", "device_structure", "结构"),
        "eqe": ("EQE", "eqe", "外量子效率"),
        "cie": ("CIE", "cie", "色度坐标"),
        "lifetime": ("寿命", "lifetime"),
        "best_eqe": ("最高EQE", "best_eqe"),
        "optimization_level": ("优化层级", "optimization_level"),
        "optimization_strategy": ("优化策略", "optimization_strategy"),
        "optimization_details": ("优化详情",),
        "key_findings": ("关键发现",),
        "eqe_source": ("EQE原文", "eqe_source"),
        "cie_source": ("CIE原文", "cie_source"),
        "lifetime_source": ("寿命原文", "lifetime_source"),
        "structure_source": ("结构原文", "structure_source"),
    }
    
    def __init__(self, output_dir: Union[str, Path]):
        """
        初始化报告生成器
        
        Args:
            output_dir: 输出目录
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_excel_report(
        self,
        results: list[dict],
        output_filename: Optional[str] = None,
        sort_by_if: bool = True,
    ) -> Path:
        """
        生成 Excel 报告
        
        Args:
            results: 提取结果列表
            output_filename: 输出文件名(可选)
            sort_by_if: 是否按影响因子排序
        
        Returns:
            生成的文件路径
        """
        # 生成文件名
        if output_filename is None:
            output_filename = self._build_default_filename("xlsx")

        output_path = self._build_unique_output_path(output_filename)
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "论文分析结果"
        
        # 样式定义
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        cell_font = Font(size=10, name="Arial")
        thin_border = Border(
            left=Side(style="thin", color="BFBFBF"),
            right=Side(style="thin", color="BFBFBF"),
            top=Side(style="thin", color="BFBFBF"),
            bottom=Side(style="thin", color="BFBFBF"),
        )
        
        # 写入表头
        for col, (header, _) in enumerate(self.REPORT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        
        # 排序
        if sort_by_if:
            results = sorted(results, key=self._sort_key_by_if, reverse=True)
        
        # 写入数据
        for row_idx, result in enumerate(results, 2):
            row_cells = []
            for col_idx, (_, field_key) in enumerate(self.REPORT_COLUMNS, 1):
                value = self._format_cell_value(result, field_key)
                cell = ws.cell(row=row_idx, column=col_idx)
                rich_value = self._build_rich_text_value(value)
                cell.value = rich_value if rich_value is not None else value
                cell.font = cell_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
                row_cells.append(cell)
                
                # 高亮有数据的关键指标列
                if field_key in {"eqe", "cie", "lifetime"} and value:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

            if self._count_abnormal_fields(result) > 3:
                light_fill = PatternFill(start_color="FFF9EF", end_color="FFF9EF", fill_type="solid")
                for cell in row_cells:
                    if cell.fill.fill_type == "solid" and cell.fill.start_color.rgb in {"00E2EFDA", "FFE2EFDA"}:
                        continue
                    cell.fill = light_fill
        
        # 设置列宽 - 按照 PRD 3.0 规范
        column_widths = {
            "file": 30,
            "url": 45,
            "journal": 28,
            "impact_factor": 16,
            "authors": 25,
            "processing_status": 34,
            "title": 50,
            "structure": 45,
            "eqe": 20,
            "cie": 20,
            "lifetime": 22,
            "best_eqe": 16,
            "optimization_level": 24,
            "optimization_strategy": 40,
            "optimization_details": 40,
            "key_findings": 40,
            "eqe_source": 50,
            "cie_source": 50,
            "lifetime_source": 50,
            "structure_source": 50,
        }

        for col_idx, (_, field_key) in enumerate(self.REPORT_COLUMNS, 1):
            width = column_widths.get(field_key, 20)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # 设置行高
        ws.row_dimensions[1].height = 35
        for row in range(2, len(results) + 2):
            ws.row_dimensions[row].height = 80
        
        # 冻结首行
        ws.freeze_panes = "A2"
        
        # 保存
        wb.save(output_path)
        logger.info(f"[Report] Excel report saved: {output_path}")
        
        return output_path
    
    def _format_cell_value(self, result: dict, field_key: str) -> str:
        """
        格式化单元格值
        
        按照 PRD 3.0 规范：
        - 多器件数据列（结构、EQE、CIE、寿命）已在 to_excel_row() 中使用 \n 拼接
        - 数据溯源列拆分为独立的原文引用列
        
        Args:
            result: 提取结果
            field_key: 规范化字段名
        
        Returns:
            格式化后的值
        """
        field_key = self._normalize_field_key(field_key)

        # 多器件数据列 - 已在 PaperData.to_excel_row() 中处理为换行拼接格式
        if field_key == "optimization_level":
            optimization = result.get("optimization", {}) if isinstance(result.get("optimization"), dict) else {}
            return (
                result.get("优化层级")
                or result.get("optimization_level")
                or optimization.get("level")
                or ""
            )

        if field_key == "optimization_strategy":
            optimization = result.get("optimization", {}) if isinstance(result.get("optimization"), dict) else {}
            return (
                result.get("优化策略")
                or result.get("optimization_strategy")
                or optimization.get("strategy")
                or ""
            )

        if field_key == "optimization_details":
            return result.get("优化详情") or ""

        if field_key == "key_findings":
            optimization = result.get("optimization", {}) if isinstance(result.get("optimization"), dict) else {}
            return result.get("关键发现") or optimization.get("key_findings") or ""

        value = self._get_mapped_value(result, field_key)
        if isinstance(value, (list, dict)):
            return json.dumps(value, ensure_ascii=False)

        if field_key == "impact_factor":
            return self._coerce_if_value(value)

        return str(value) if value not in (None, "") else ""
    
    def generate_json_report(
        self,
        results: list[dict],
        output_filename: Optional[str] = None,
        sort_by_if: bool = True,
    ) -> Path:
        """
        生成 JSON 报告
        
        Args:
            results: 提取结果列表
            output_filename: 输出文件名
            sort_by_if: 是否按影响因子排序
        
        Returns:
            生成的文件路径
        """
        if output_filename is None:
            output_filename = self._build_default_filename("json")

        output_path = self._build_unique_output_path(output_filename)
        
        if sort_by_if:
            results = sorted(results, key=self._sort_key_by_if, reverse=True)
        
        with output_path.open("w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)

        logger.info(f"[Report] JSON report saved: {output_path}")
        
        return output_path
    
    def generate_error_log(
        self,
        errors: list[dict],
        output_filename: Optional[str] = None,
    ) -> Optional[Path]:
        """
        生成错误日志
        
        Args:
            errors: 错误列表
            output_filename: 输出文件名
        
        Returns:
            生成的文件路径(如果有错误)
        """
        if not errors:
            return None
        
        if output_filename is None:
            output_filename = "error_log.txt"
        
        output_path = self.output_dir / output_filename
        
        with output_path.open("w", encoding="utf-8") as f:
            f.write("=" * 70 + "\n")
            f.write("PaperInsight Error Log\n")
            f.write(f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total errors: {len(errors)}\n")
            f.write("=" * 70 + "\n\n")
            
            for idx, error in enumerate(errors, 1):
                f.write(f"[Error {idx}]\n")
                f.write(f"Time: {error.get('timestamp', 'N/A')}\n")
                f.write(f"File: {error.get('pdf_name', 'N/A')}\n")
                f.write(f"Type: {error.get('error_type', 'N/A')}\n")
                f.write(f"Message: {error.get('error_message', 'N/A')}\n")
                if error.get('context'):
                    f.write(f"Context: {error['context']}\n")
                f.write("-" * 70 + "\n\n")
        
        logger.info(f"[ErrorLog] saved: {output_path}")

        return output_path

    def _get_mapped_value(self, result: dict, field_key: str):
        for key in self.FIELD_MAPPING.get(field_key, (field_key,)):
            if key in result and result[key] not in (None, ""):
                return result[key]
        return ""

    def _sort_key_by_if(self, result: dict) -> float:
        return self._coerce_if_value(self._get_mapped_value(result, "impact_factor"))

    def _build_default_filename(self, extension: str) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"论文分析报告_{timestamp}"
        return f"{base_name}.{extension}"

    def _build_unique_output_path(self, output_filename: str) -> Path:
        output_path = self.output_dir / output_filename
        if not output_path.exists():
            return output_path

        stem = output_path.stem
        suffix = output_path.suffix
        counter = 1
        while True:
            candidate = self.output_dir / f"{stem}_{counter}{suffix}"
            if not candidate.exists():
                return candidate
            counter += 1

    @staticmethod
    def _coerce_if_value(value) -> float:
        if isinstance(value, (int, float)):
            return float(value)

        if isinstance(value, str):
            match = __import__("re").search(r"([0-9]+(?:\.[0-9]+)?)", value)
            if match:
                return float(match.group(1))

        return 0.0

    def _normalize_field_key(self, field_key: str) -> str:
        if field_key in self.FIELD_MAPPING or field_key in dict(self.REPORT_COLUMNS):
            return field_key

        for canonical_key, aliases in self.FIELD_MAPPING.items():
            if field_key in aliases:
                return canonical_key

        return field_key

    @staticmethod
    def _build_rich_text_value(value):
        if not isinstance(value, str):
            return None

        lines = [line.strip() for line in value.splitlines() if line.strip()]
        if len(lines) != 2:
            return None

        if not lines[0].startswith("中文：") or not lines[1].startswith("English:"):
            return None

        return CellRichText(
            TextBlock(InlineFont(b=True), lines[0]),
            "\n",
            lines[1],
        )

    def _count_abnormal_fields(self, result: dict) -> int:
        important_keys = [
            "journal",
            "impact_factor",
            "title",
            "structure",
            "eqe",
            "cie",
            "lifetime",
            "best_eqe",
            "optimization_strategy",
        ]
        count = 0
        for key in important_keys:
            value = self._format_cell_value(result, key)
            if value in ("", None, 0, 0.0):
                count += 1
        return count
