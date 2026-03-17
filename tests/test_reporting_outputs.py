import openpyxl

from paperinsight.core.reporter import ReportGenerator
from paperinsight.models.schemas import PaperData, PaperInfo


def test_report_generator_uses_bilingual_headers_and_unique_filename(tmp_path):
    reporter = ReportGenerator(tmp_path)
    results = [
        {
            "File": "paper.pdf",
            "URL": "file:///tmp/paper.pdf",
            "期刊": "Nature",
            "原始期刊标题": "NATURE",
            "原始ISSN": "1476-4687",
            "原始eISSN": "1476-4687",
            "匹配期刊": "Nature",
            "匹配ISSN": "1476-4687",
            "匹配方式": "issn",
            "期刊主页": "https://example.test/journal/nature",
            "影响因子": 12.3,
            "影响因子年份": 2025,
            "影响因子来源": "MJL_WEB",
            "影响因子状态": "OK",
            "作者": "Alice",
            "标题": "中文：示例标题\nEnglish: Sample Title",
            "器件结构": "ITO/EML/Al",
            "EQE": "20.5%",
            "CIE": "(0.21, 0.32)",
            "寿命": "120 h",
            "最高EQE": "20.5%",
            "优化层级": "界面工程",
            "优化策略": "中文：中文总结\nEnglish: English summary",
        }
    ]

    first_path = reporter.generate_excel_report(results)
    second_path = reporter.generate_excel_report(results)

    assert first_path.name.startswith("论文分析报告_")
    assert second_path.name.startswith("论文分析报告_")
    assert first_path != second_path

    workbook = openpyxl.load_workbook(first_path)
    sheet = workbook.active
    headers = [sheet.cell(row=1, column=idx).value for idx in range(1, 16)]
    values = [sheet.cell(row=2, column=idx).value for idx in range(1, 16)]

    assert headers == [
        "文件名 File",
        "文件地址 URL",
        "期刊名称 Journal",
        "原始期刊标题 Raw Journal",
        "原始ISSN Raw ISSN",
        "原始eISSN Raw eISSN",
        "匹配期刊 Matched Journal",
        "匹配ISSN Matched ISSN",
        "匹配方式 Match Method",
        "期刊主页 Journal Profile URL",
        "影响因子 Impact Factor",
        "影响因子年份 IF Year",
        "影响因子来源 IF Source",
        "影响因子状态 IF Status",
        "作者 Authors",
    ]
    assert values[:10] == [
        "paper.pdf",
        "file:///tmp/paper.pdf",
        "Nature",
        "NATURE",
        "1476-4687",
        "1476-4687",
        "Nature",
        "1476-4687",
        "issn",
        "https://example.test/journal/nature",
    ]
    assert values[10:15] == [12.3, "2025", "MJL_WEB", "OK", "Alice"]


def test_paper_data_to_excel_row_includes_journal_enrichment_fields():
    paper_data = PaperData(
        paper_info=PaperInfo(
            journal_name="Nature",
            raw_journal_title="NATURE",
            raw_issn="1476-4687",
            raw_eissn="1476-4687",
            matched_journal_title="Nature",
            matched_issn="1476-4687",
            match_method="issn",
            journal_profile_url="https://example.test/journal/nature",
            impact_factor=12.3,
            impact_factor_year=2025,
            impact_factor_source="MJL_WEB",
            impact_factor_status="OK",
        )
    )

    row = paper_data.to_excel_row()

    assert row["期刊"] == "Nature"
    assert row["原始期刊标题"] == "NATURE"
    assert row["原始ISSN"] == "1476-4687"
    assert row["原始eISSN"] == "1476-4687"
    assert row["匹配期刊"] == "Nature"
    assert row["匹配ISSN"] == "1476-4687"
    assert row["匹配方式"] == "issn"
    assert row["期刊主页"] == "https://example.test/journal/nature"
    assert row["影响因子年份"] == 2025
    assert row["影响因子来源"] == "MJL_WEB"
    assert row["影响因子状态"] == "OK"


def test_report_generator_highlights_rows_with_many_missing_fields(tmp_path):
    reporter = ReportGenerator(tmp_path)
    results = [
        {
            "File": "bad.pdf",
            "URL": "file:///tmp/bad.pdf",
            "处理结果/简述": "处理失败：数据提取 - 返回空结果",
            "标题": "",
            "期刊": "",
            "影响因子": "",
            "作者": "",
            "器件结构": "",
            "EQE": "",
            "CIE": "",
            "寿命": "",
            "最高EQE": "",
            "优化策略": "",
        }
    ]

    path = reporter.generate_excel_report(results)
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    assert sheet["A2"].fill.start_color.rgb in {"00FFF9EF", "FFFFF9EF"}
